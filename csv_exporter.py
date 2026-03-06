import csv
import os
from datetime import datetime

COLUMNS = [
    ("full_name", "Full Name"),
    ("job_title", "Job Title"),
    ("company", "Company"),
    ("email", "Email"),
    ("email_status", "Email Status"),
    ("linkedin_url", "LinkedIn URL"),
    ("category", "Category"),
    ("source", "Source"),
]


def export_to_csv(leads: list, output_path: str) -> str:
    if not output_path.endswith(".csv"):
        output_path += ".csv"

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        headers = [display for _, display in COLUMNS] + ["Export Date"]
        writer.writerow(headers)

        for lead in leads:
            row = [lead.get(key, "") for key, _ in COLUMNS] + [timestamp]
            writer.writerow(row)

    return os.path.abspath(output_path)


def export_to_excel(leads: list, output_path: str) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("    ⚠ openpyxl not installed. Run: pip install openpyxl")
        return None

    if not output_path.endswith(".xlsx"):
        output_path += ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        bottom=Side(style="medium", color="1F3864"),
        right=Side(style="thin", color="D6DCE4"),
    )

    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=False)
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    valid_font = Font(name="Calibri", size=10, color="006100")
    valid_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    headers = [display for _, display in COLUMNS] + ["Export Date"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    for row_idx, lead in enumerate(leads, 2):
        row_data = [lead.get(key, "") for key, _ in COLUMNS] + [timestamp]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment

            if row_idx % 2 == 0:
                cell.fill = alt_fill

            if col_idx == 4 and value:
                email_status = lead.get("email_status", "")
                if email_status == "valid":
                    cell.font = valid_font
                    cell.fill = valid_fill

    column_widths = {
        1: 22,
        2: 30,
        3: 22,
        4: 30,
        5: 14,
        6: 45,
        7: 16,
        8: 14,
        9: 20,
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{len(leads) + 1}"

    wb.save(output_path)
    return os.path.abspath(output_path)


def export_leads(leads: list, output_path: str, format: str = "both") -> dict:
    base_path = output_path
    for ext in [".csv", ".xlsx", ".xls"]:
        if base_path.endswith(ext):
            base_path = base_path[:-len(ext)]
            break

    result = {}

    if format in ("csv", "both"):
        csv_path = export_to_csv(leads, base_path + ".csv")
        if csv_path:
            result["csv"] = csv_path

    if format in ("excel", "both"):
        xlsx_path = export_to_excel(leads, base_path + ".xlsx")
        if xlsx_path:
            result["excel"] = xlsx_path

    return result